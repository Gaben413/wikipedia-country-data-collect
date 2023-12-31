from bs4 import BeautifulSoup
import requests
import json
import csv
from openpyxl import Workbook
import sys

def FetchCountriesLinks():
    URL = 'https://en.wikipedia.org/wiki/List_of_sovereign_states_and_dependent_territories_in_Asia'
    page = requests.get(URL)
    soup = BeautifulSoup(page.text, 'html.parser')
    tables = soup.find_all("table")

    exclusion_countries = [
        'Turkey',
        'Bahrain',
        'Kuwait',
        'Oman',
        'Qatar',
        'Saudi Arabia',
        'United Arab Emirates',
        'Yemen',
        'Abkhazia',
        'Armenia',
        'Artsakh',
        'Azerbaijan',
        'South Ossetia',
        'Iraq',
        'Israel',
        'Jordan',
        'Lebanon',
        'Palestine',
        'Syria',
        'Iran',
        'Akrotiri and Dhekelia',
        'Cyprus',
        'Northern Cyprus ',
        'Egypt',
        'Russia',
    ]
    links_dict = []

    targets = [1, 2, 5]

    for target in targets:
        for row in tables[target].tbody.find_all('tr'):
            columns = row.find_all('td')

            if(columns != []):
                if '#' not in columns[2].find("a").get("href"):
                    #print(f'LINK: {columns[2].find("a").get("title")} | COUNTRY: {columns[2].find("a").get("href")}')
                    if(columns[2].find('a').get('title') not in exclusion_countries):
                        links_dict.append({
                            'country': columns[2].find("a").get("title"),
                            'link': columns[2].find("a").get("href"),
                        })

    return links_dict

dicts = FetchCountriesLinks()

wikiURL = 'https://en.wikipedia.org'

def GetGDP(input):
    parts = input.replace('\xa0',' ').replace('[',' [').replace('$', '$ ').replace(',','.').split(' ')

    for part in parts:
        if(part.replace('.','').isdigit()):
            return part
        
def GetReligion(input):
    try:
        if(len(input.find_all('li')) > 0):
            return input.find_all('li')[0].find_all('a')[0].text
        else:
            return input.find_all('a')[0].text
    except:
        print("!ATTENTTION!     RELIGION NOT FOUND      !ATTENTTION!")
        return 'Unknown'
    
def GetLanguage(input):
    if(len(input.find_all('li')) > 0):
        if('[' in input.find_all('li')[0].find('a').text):
            return input.find_all('li')[0].text
        else:
            return input.find_all('li')[0].find('a').text
    else:
        return input.find_all('a')[0].text

def GetYear(input):
    parts = input.find('div').text.replace('\xa0', ' ').split(' ')
    
    output = 0

    for part in parts:
        if(part.isdigit()):
            output = part
            
    return output

def FetchPageData(input):
    page = requests.get(wikiURL+input['link'])
    soup = BeautifulSoup(page.text, 'html.parser')
    tables = soup.find_all("table", class_='infobox')[0]

    #print(tables)

    trs = tables.tbody.find_all('tr')
    #print(len(trs))
    print(f"Fetching Data from {input['country']}")

    obj = {
        'pops': 0,
        'rel': 0,
        'gdp': 0,
        'lang': 0,
        'hdi': 0
    }

    for i,tr in enumerate(trs):
        if(tr.th != None):
            if('Population' in tr.th.text):
                #print(f'Population Tab AT: {i}')
                obj['pops'] = i

            if('Religion' in tr.th.text):
                #print(f'Religion Tab AT: {i}')
                obj['rel'] = i

            if(tr.th.find('a') != None and 'GDP' in tr.th.find('a').text):
                #print(f'Religion Tab AT: {i}')
                obj['gdp'] = i

            if('language' in tr.th.text and obj['lang'] == 0):
                #print(f'Religion Tab AT: {i}')
                obj['lang'] = i

            if('HDI' in tr.th.text):
                #print(f'Religion Tab AT: {i}')
                obj['hdi'] = i
    '''
    print(obj)
    #POPS:
    print(trs[(obj['pops']+1)].th.find('div').text.replace('\xa0', ' ').split(' ')[1])
    print(trs[(obj['pops']+1)].td.text.replace('[',' [').strip().split(' ')[0].replace(',',''))

    #GDP
    print(trs[(obj['gdp'])].td.text.replace('\xa0',' ').split(' ')[0])
    print(GetGDP(trs[(obj['gdp']+1)].td.text))

    #RELIGION
    print(GetReligion(trs[obj['rel']].td))

    #LANGUAGE
    print(GetLanguage(trs[obj['lang']].td))

    #HDI
    print(trs[obj['hdi']].th.find('span').text.replace('(','').replace(')',''))
    print(float(trs[obj['hdi']].td.text.replace('[',' [').replace(']', '] ').strip().split(' ')[0]))
    '''    

    hdi = ''
    hdi_year = ''

    try:
        hdi = float(trs[obj['hdi']].td.text.replace('[',' [').replace(']', '] ').strip().split(' ')[0])
    except:
        hdi = 'Unknown'

    try:
        hdi_year = int(trs[obj['hdi']].th.find('span').text.replace('(','').replace(')',''))
    except:
        hdi_year = 'Unknown'

    data = {
        'country': input['country'],
        'population': int(trs[(obj['pops']+1)].td.text.replace('[',' [').strip().split(' ')[0].replace(',','')),
        'pops_year': int(GetYear(trs[(obj['pops']+1)].th)),
        'GDP_year': int(trs[(obj['gdp'])].td.text.replace('\xa0',' ').split(' ')[0]),
        'GDP_total': float(GetGDP(trs[(obj['gdp']+1)].td.text)),
        'religion': GetReligion(trs[obj['rel']].td),
        'language': GetLanguage(trs[obj['lang']].td),
        'hdi_year': hdi_year,
        'hdi': hdi
    }
    
    #print(data)

    return data

data_array = []

for i,dict in enumerate(dicts):
    '''
    sys.stdout.write("\r" + str(int((i/len(dicts))*100))+"%")
    sys.stdout.flush()
    '''
    print("\r" + str(int((i/len(dicts))*100))+"%")
    data_array.append(FetchPageData(dict))

#print(len(data_array))

def saveJSON(input):
    json_dump = json.dumps(input, indent=3, ensure_ascii=False)

    #print(json_dump)

    with open('data.json', 'w', encoding='utf8') as outfile:
        outfile.write(json_dump)
    
    print('JSON saved')

headers = [
    'country',
    'population',
    'pops_year',
    'GDP_year',
    'GDP_total',
    'religion',
    'language',
    'hdi_year',
    'hdi'
]

def saveCSV(input):
    with open('data.csv', mode='w', encoding='utf8') as outfile:

        filewriter = csv.writer(outfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        filewriter.writerow(headers)

        for obj in input:

            filewriter.writerow([
                obj['country'],
                obj['population'],
                obj['pops_year'],
                obj['GDP_year'],
                obj['GDP_total'],
                obj['religion'],
                obj['language'],
                obj['hdi_year'],
                obj['hdi'],
            ])
    print('CSV saved')

def saveEXCEL(input):

    filename = 'data.xlsx'

    workbook = Workbook()
    sheet = workbook.active

    #SET HEADERS
    for i, header in enumerate(headers):
        sheet.cell(row=1, column=(i+1)).value = header

    #ADD DATA
    for i, obj in enumerate(input):
        
        sheet.cell(row=(i+2), column=1).value = obj['country']
        sheet.cell(row=(i+2), column=2).value = obj['population']
        sheet.cell(row=(i+2), column=3).value = obj['pops_year']
        sheet.cell(row=(i+2), column=4).value = obj['GDP_year']
        sheet.cell(row=(i+2), column=5).value = obj['GDP_total']
        sheet.cell(row=(i+2), column=6).value = obj['religion']
        sheet.cell(row=(i+2), column=7).value = obj['language']
        sheet.cell(row=(i+2), column=8).value = obj['hdi_year']
        sheet.cell(row=(i+2), column=9).value = obj['hdi']

    sheet.freeze_panes = "A2"        

    workbook.save(filename=filename)

    print('EXCEL saved')

saveJSON(data_array)
saveCSV(data_array)
saveEXCEL(data_array)